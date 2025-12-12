import fitz  # PyMuPDF
import re
import pandas as pd
import os
from pathlib import Path


def extract_equipment_data(pdf_path):
    """
    Extract equipment data from electrical one line diagrams ONLY
    Excludes right side boxes with legends, notes, and company info
    Returns data organized left to right, top to bottom
    Includes Primary From and Alternate From information
    """
    # Check if file exists
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    doc = fitz.open(pdf_path)
    all_equipment = []
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        page_width = page.rect.width
        page_height = page.rect.height
        
        # Define diagram area (exclude right side boxes)
        diagram_right_boundary = page_width * 0.75
        diagram_top_boundary = page_height * 0.05
        diagram_bottom_boundary = page_height * 0.95
        
        # Extract text with positions
        blocks = page.get_text("dict")["blocks"]
        
        # Collect all text elements with positions (only from diagram area)
        elements = []
        for block in blocks:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"].strip()
                        x_pos = span["bbox"][0]
                        y_pos = span["bbox"][1]
                        
                        # Only include text from the diagram area (left side)
                        if (
                            text
                            and len(text) > 0
                            and x_pos < diagram_right_boundary
                            and y_pos > diagram_top_boundary
                            and y_pos < diagram_bottom_boundary
                        ):
                            elements.append(
                                {
                                    "text": text,
                                    "x0": span["bbox"][0],
                                    "y0": span["bbox"][1],
                                    "x1": span["bbox"][2],
                                    "y1": span["bbox"][3],
                                    "font_size": span["size"],
                                }
                            )
        
        # Sort top to bottom, left to right
        elements.sort(key=lambda e: (round(e["y0"] / 15) * 15, e["x0"]))
        
        # Equipment name pattern (5 letters + 3 digits)
        equipment_pattern = re.compile(r"\b([A-Z]{5}\d{3})\b")
        
        # Additional filters to exclude non diagram equipment references
        exclude_keywords = [
            "LINETYPE",
            "LEGEND",
            "NOTES",
            "KEYED",
            "ISSUES",
            "REVISIONS",
            "PROPRIETARY",
            "CONSULTING",
            "ENGINEERS",
            "ARCHITECTS",
            "ONE-LINE GENERAL NOTES",
            "EDGECONNEX",
            "ASSET CODE",
            "IDENTIFICATION SYSTEM",
            "BURR COMPUTER",
        ]
        
        # Build a lookup dictionary for quick equipment name search
        equipment_positions = {}
        for i, element in enumerate(elements):
            match = equipment_pattern.search(element["text"])
            if match:
                equip_name = match.group(1)
                equipment_positions[equip_name] = {
                    "index": i,
                    "x0": element["x0"],
                    "y0": element["y0"],
                    "x1": element["x1"],
                    "y1": element["y1"],
                }
        
        # Process each element
        for i, element in enumerate(elements):
            text = element["text"]
            match = equipment_pattern.search(text)
            
            if match:
                equipment_name = match.group(1)
                
                # Skip if this appears to be in a legend or note box
                is_in_legend = False
                for j in range(max(0, i - 3), min(i + 3, len(elements))):
                    context_text = elements[j]["text"].upper()
                    if any(keyword in context_text for keyword in exclude_keywords):
                        is_in_legend = True
                        break
                
                if is_in_legend:
                    continue
                
                equipment_type = equipment_name[:3]  # First 3 letters
                
                # Collect properties (text below equipment)
                properties = []
                base_y = element["y0"]
                base_x = element["x0"]
                
                # Common property patterns to identify
                property_patterns = [
                    r"\d+A",  # Amperage
                    r"\d+kW",  # Kilowatts
                    r"\d+kVA",  # Kilovolt ampere
                    r"\d+V",  # Voltage
                    r"\d+Y/\d+V",  # Voltage configuration
                    r"\d+⌀",  # Phase
                    r"\d+W",  # Wire
                    r"\d+kAIC",  # AIC rating
                    r"\d+AF|AT",  # Frame/Trip rating
                    r"\d+%",  # Percentage
                    r"K\d+",  # K-rating
                    r"NEMA\s+\d+",  # NEMA rating
                    r"\(ROOF\)",  # Location
                    r"\([A-Z\s]+ROOM\)",  # Room location
                    r"\([A-Z\s]+GALLERY\s*\d*\)",  # Gallery location
                    r"N\.C\.|N\.O\.",  # Normally closed/open
                    r"LSIG|ERMS|LSI|S\.T\.U\.",  # Trip settings
                    r"POLE",  # Pole configuration
                ]
                
                # Look at nearby text (within 80 pixels below and 200 pixels horizontally)
                for j in range(i + 1, min(i + 15, len(elements))):
                    next_elem = elements[j]
                    
                    # Check proximity
                    y_diff = next_elem["y0"] - base_y
                    x_diff = abs(next_elem["x0"] - base_x)
                    
                    if y_diff > 80:  # Too far below
                        break
                    
                    if y_diff < 0 or x_diff > 200:  # Above or too far horizontally
                        continue
                    
                    # Stop at next equipment
                    if equipment_pattern.search(next_elem["text"]):
                        break
                    
                    # Check if it matches property patterns
                    prop_text = next_elem["text"]
                    is_property = False
                    
                    for pattern in property_patterns:
                        if re.search(pattern, prop_text, re.IGNORECASE):
                            is_property = True
                            break
                    
                    # Also include specific keywords
                    keywords = [
                        "STATIC",
                        "UPS",
                        "SYSTEM",
                        "GENERATOR",
                        "SWITCHBOARD",
                        "PANEL",
                        "INVERTER",
                        "RECTIFIER",
                        "INPUT",
                        "OUTPUT",
                        "MAINTENANCE",
                        "BYPASS",
                        "CABINET",
                        "MECHANICAL",
                        "DIST",
                        "TRANSFORMER",
                        "BREAKER",
                        "VFD",
                        "RATED",
                        "CHILLER",
                        "BATTERY",
                        "STRING",
                        "FUSES",
                        "SWITCH",
                        "DISTRIBUTION",
                        "BUSWAY",
                        "RESERVE",
                        "HUMIDIFIER",
                        "AHU",
                        "FREIGHT",
                    ]
                    
                    if any(keyword in prop_text.upper() for keyword in keywords):
                        is_property = True
                    
                    # Skip common non property text
                    skip_patterns = [
                        "RE:",
                        "O.F.C.I.",
                        "E103",
                        "E104",
                        "E106",
                        "E107",
                        "E108",
                        "E110",
                        "E114",
                        "E115",
                        "E118",
                        "E119",
                        "E121",
                        "E126",
                        "E127",
                        "E157",
                        "SPD",
                        "PM1",
                        "PM2",
                        "METER",
                        "GFPE",
                        "ATS PLC",
                        "BY",
                        "MANUF",
                    ]
                    if any(skip in prop_text for skip in skip_patterns):
                        is_property = False
                    
                    # Skip if property text contains excluded keywords
                    if any(keyword in prop_text.upper() for keyword in exclude_keywords):
                        is_property = False
                    
                    if is_property and prop_text not in properties:
                        properties.append(prop_text)
                
                # Extract "Primary From" and "Alternate From"
                primary_from_list = []
                alternate_from_list = []
                
                # Search area above the equipment
                search_y_max = base_y
                search_y_min = base_y - 150
                search_x_min = base_x - 300
                search_x_max = element["x1"] + 300
                
                # Find all equipment names in the search area above
                above_equipment = []
                for j, elem in enumerate(elements):
                    if j == i:
                        continue
                    
                    elem_y = elem["y0"]
                    elem_x = elem["x0"]
                    
                    # Check if element is in the search area (above)
                    if (
                        elem_y >= search_y_min
                        and elem_y < search_y_max
                        and elem_x >= search_x_min
                        and elem_x <= search_x_max
                    ):
                        # Check if this element contains an equipment name
                        elem_match = equipment_pattern.search(elem["text"])
                        if elem_match:
                            found_equip = elem_match.group(1)
                            
                            # Skip if it is in the legend area
                            skip_this = False
                            for k in range(max(0, j - 3), min(j + 3, len(elements))):
                                ctx_text = elements[k]["text"].upper()
                                if any(keyword in ctx_text for keyword in exclude_keywords):
                                    skip_this = True
                                    break
                            
                            if not skip_this and found_equip != equipment_name:
                                # Calculate vertical distance
                                v_distance = elem_y - base_y
                                above_equipment.append(
                                    {
                                        "name": found_equip,
                                        "y_distance": v_distance,
                                        "x": elem_x,
                                        "y": elem_y,
                                    }
                                )
                
                # Sort by vertical distance (closest first, least negative)
                above_equipment.sort(key=lambda e: -e["y_distance"])
                
                # Classify as Primary From or Alternate From
                if len(above_equipment) == 1:
                    primary_from_list.append(above_equipment[0]["name"])
                elif len(above_equipment) > 1:
                    # Group by similar y-distance (within 30 pixels)
                    y_groups = []
                    current_group = [above_equipment[0]]
                    
                    for k in range(1, len(above_equipment)):
                        if (
                            abs(
                                above_equipment[k]["y_distance"]
                                - above_equipment[k - 1]["y_distance"]
                            )
                            < 30
                        ):
                            current_group.append(above_equipment[k])
                        else:
                            y_groups.append(current_group)
                            current_group = [above_equipment[k]]
                    y_groups.append(current_group)
                    
                    # First group (closest) contains primary candidates
                    closest_group = y_groups[0]
                    
                    if len(closest_group) == 1:
                        primary_from_list.append(closest_group[0]["name"])
                    else:
                        # Multiple equipment at similar height
                        # Choose the one closest horizontally as primary
                        closest_group.sort(key=lambda e: abs(e["x"] - base_x))
                        primary_from_list.append(closest_group[0]["name"])
                        
                        # Rest are alternates
                        for eq in closest_group[1:]:
                            alternate_from_list.append(eq["name"])
                    
                    # All other groups are alternates
                    for group in y_groups[1:]:
                        for eq in group:
                            alternate_from_list.append(eq["name"])
                
                primary_from = ", ".join(primary_from_list) if primary_from_list else "-"
                alternate_from = (
                    ", ".join(alternate_from_list) if alternate_from_list else "-"
                )
                
                all_equipment.append(
                    {
                        "Page": page_num + 1,
                        "Equipment": equipment_name,
                        "Type": equipment_type,
                        "Primary From": primary_from,
                        "Alternate From": alternate_from,
                        "Properties": ", ".join(properties) if properties else "-",
                        "_sort_y": round(base_y / 15) * 15,
                        "_sort_x": base_x,
                    }
                )
    
    # Final sort. left to right, top to bottom
    all_equipment.sort(key=lambda e: (e["Page"], e["_sort_y"], e["_sort_x"]))
    
    # Remove helper sort keys before returning
    for equipment in all_equipment:
        equipment.pop("_sort_y", None)
        equipment.pop("_sort_x", None)
    
    return all_equipment


def run_type2_extractor(pdf_path: str):
    """
    Wrapper used by the FastAPI endpoint for Type 2 PDFs.
    Takes a single uploaded PDF path.
    Returns.
      df. pandas DataFrame
      equipment_list. list of dicts with keys expected by the UI
    """
    equipment_data = extract_equipment_data(pdf_path)
    
    # Build DataFrame
    if equipment_data:
        df = pd.DataFrame(equipment_data)
        # keep a clean column order if columns exist
        desired_cols = [
            "Page",
            "Equipment",
            "Type",
            "Primary From",
            "Alternate From",
            "Properties",
        ]
        df = df[[c for c in desired_cols if c in df.columns]]
    else:
        df = pd.DataFrame(
            columns=[
                "Page",
                "Equipment",
                "Type",
                "Primary From",
                "Alternate From",
                "Properties",
            ]
        )
    
    # Build list of dicts for the frontend table
    equipment_list = []
    for row in equipment_data:
        equipment_list.append(
            {
                "Equipment": row.get("Equipment", ""),
                "Type": row.get("Type", ""),
                "Properties": row.get("Properties", ""),
                "Primary From": row.get("Primary From", ""),
                "Alternate From": row.get("Alternate From", ""),
            }
        )
    
    return df, equipment_list


def save_to_excel(all_data, output_path):
    """
    Save all extracted data to Excel with multiple sheets
    One combined sheet and separate sheets for each system
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Create combined sheet with all systems
        all_combined = []
        for system_name, equipment_list in all_data.items():
            for equipment in equipment_list:
                equipment_with_system = {"System": system_name}
                equipment_with_system.update(equipment)
                all_combined.append(equipment_with_system)
        
        if all_combined:
            df_combined = pd.DataFrame(all_combined)
            # Reorder columns
            column_order = [
                "System",
                "Page",
                "Equipment",
                "Type",
                "Primary From",
                "Alternate From",
                "Properties",
            ]
            df_combined = df_combined[column_order]
            df_combined.to_excel(writer, index=False, sheet_name="All Systems")
            worksheet = writer.sheets["All Systems"]
            format_worksheet(worksheet, has_system_column=True)
        
        # Create separate sheets for each system
        for system_name, equipment_list in all_data.items():
            df = pd.DataFrame(equipment_list)
            if not df.empty:
                # Reorder columns
                column_order = [
                    "Page",
                    "Equipment",
                    "Type",
                    "Primary From",
                    "Alternate From",
                    "Properties",
                ]
                df = df[column_order]
                df.to_excel(writer, index=False, sheet_name=system_name)
                worksheet = writer.sheets[system_name]
                format_worksheet(worksheet, has_system_column=False)
    
    total_equipment = sum(len(eq_list) for eq_list in all_data.values())
    print(f"✓ Data successfully saved to. {output_path}")
    print(f"✓ Total equipment extracted. {total_equipment}")
    for system_name, equipment_list in all_data.items():
        print(f"  - {system_name}. {len(equipment_list)} items")


def format_worksheet(worksheet, has_system_column=False):
    """Format worksheet with proper styling"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    if has_system_column:
        # Adjust column widths for combined sheet
        worksheet.column_dimensions["A"].width = 15  # System
        worksheet.column_dimensions["B"].width = 10  # Page
        worksheet.column_dimensions["C"].width = 15  # Equipment
        worksheet.column_dimensions["D"].width = 10  # Type
        worksheet.column_dimensions["E"].width = 20  # Primary From
        worksheet.column_dimensions["F"].width = 20  # Alternate From
        worksheet.column_dimensions["G"].width = 80  # Properties
    else:
        # Adjust column widths for individual system sheets
        worksheet.column_dimensions["A"].width = 10  # Page
        worksheet.column_dimensions["B"].width = 15  # Equipment
        worksheet.column_dimensions["C"].width = 10  # Type
        worksheet.column_dimensions["D"].width = 20  # Primary From
        worksheet.column_dimensions["E"].width = 20  # Alternate From
        worksheet.column_dimensions["F"].width = 80  # Properties
    
    # Format header row
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Center align all columns except Properties
    if has_system_column:
        # Center align System, Page, Equipment, Type, Primary From, Alternate From columns
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=1, max_col=6
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        
        # Wrap text in Properties column (column G)
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=7, max_col=7
        ):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    else:
        # Center align Page, Equipment, Type, Primary From, Alternate From columns
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        
        # Wrap text in Properties column (column F)
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6
        ):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


# Main execution for batch local use
if __name__ == "__main__":
    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # PDF file names
    pdf_files = [
        "SYSTEM A Labels.pdf",
        "SYSTEM B Labels.pdf",
        "SYSTEM C.pdf",
        "SYSTEM D.pdf",
        "SYSTEM E.pdf",
        "SYSTEM F.pdf",
    ]
    
    # Output Excel file
    output_excel = os.path.join(script_dir, "equipment_data_all_systems.xlsx")
    
    print("=" * 80)
    print("EQUIPMENT DATA EXTRACTION FROM ELECTRICAL ONE LINE DIAGRAMS")
    print("=" * 80)
    print(f"\nScript location. {script_dir}")
    print(f"\nLooking for PDFs.")
    for pdf in pdf_files:
        print(f"  - {pdf}")
    
    # Check if files exist
    missing_files = []
    for pdf_file in pdf_files:
        pdf_path = os.path.join(script_dir, pdf_file)
        if not os.path.exists(pdf_path):
            missing_files.append(pdf_file)
    
    if missing_files:
        print("\n" + "!" * 80)
        print("ERROR. PDF FILE(S) NOT FOUND.")
        print("!" * 80)
        print("\nMissing files.")
        for pdf_file in missing_files:
            print(f"  - {pdf_file}")
        
        # List PDF files in the directory
        try:
            found_pdfs = [
                f for f in os.listdir(script_dir) if f.lower().endswith(".pdf")
            ]
            if found_pdfs:
                print(f"\nPDF files found in the script directory.")
                for pdf_file in found_pdfs:
                    print(f"  - {pdf_file}")
            else:
                print("\nNo PDF files found in the script directory.")
        except Exception:
            pass
        
        print("\n" + "!" * 80)
        exit(1)
    
    print(f"\n✓ All PDF files found.")
    print("\nExtracting equipment data from DIAGRAM AREA ONLY.")
    print("(Including Primary From and Alternate From connections)\n")
    
    try:
        # Dictionary to store data from all systems
        all_systems_data = {}
        
        # Extract data from each PDF
        for pdf_file in pdf_files:
            pdf_path = os.path.join(script_dir, pdf_file)
            system_name = pdf_file.replace(" Labels.pdf", "").replace(".pdf", "")
            
            print(f"Processing. {pdf_file}.")
            equipment_data = extract_equipment_data(pdf_path)
            all_systems_data[system_name] = equipment_data
            print(f"  ✓ Extracted {len(equipment_data)} items from {system_name}\n")
        
        # Display preview
        print("=" * 80)
        print("PREVIEW OF EXTRACTED DATA")
        print("=" * 80)
        
        for system_name, equipment_list in all_systems_data.items():
            print(f"\n{system_name}.")
            print("-" * 80)
            if equipment_list:
                df = pd.DataFrame(equipment_list)
                print(df.head(10).to_string(index=False))
                if len(equipment_list) > 10:
                    print(f". (showing first 10 of {len(equipment_list)} items)")
            else:
                print("  No equipment data found")
        
        print("\n" + "=" * 80)
        
        # Save to Excel
        save_to_excel(all_systems_data, output_excel)
        
        print("\n" + "=" * 80)
        print("EXTRACTION COMPLETE.")
        print("=" * 80)
        print(f"\nOutput file. {output_excel}")
        print("\nExcel file contains.")
        print("  - 'All Systems' sheet. Combined data from all PDFs")
        for system_name in all_systems_data.keys():
            print(f"  - '{system_name}' sheet. Individual system data")
        print("\nColumns extracted.")
        print("  • Page")
        print("  • Equipment")
        print("  • Type")
        print("  • Primary From (equipment feeding this one)")
        print("  • Alternate From (alternate feed sources)")
        print("  • Properties")
        
    except Exception as e:
        print("\n" + "!" * 80)
        print("ERROR DURING EXTRACTION.")
        print("!" * 80)
        print(f"\nError details. {str(e)}")
        import traceback
        
        print("\nFull traceback.")
        traceback.print_exc()
        print("\n" + "!" * 80)
        exit(1)
