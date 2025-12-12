from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles

from pathlib import Path
import uuid
import time
import shutil

from typing import List, Dict, Any
from pydantic import BaseModel

from extract_equipment_simple import main as run_type1_extractor
from Extract_equipment_simple2 import run_type2_extractor, save_to_excel

# --------------------------------------------------------------------
# Paths / app setup
# --------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
UPLOADS_DIR = BASE_DIR / "uploads"
OUTPUTS_DIR = BASE_DIR / "outputs"

UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/outputs", StaticFiles(directory=OUTPUTS_DIR), name="outputs")


@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the main HTML page"""
    index_path = STATIC_DIR / "index.html"
    return index_path.read_text(encoding="utf-8")


def _save_upload_to_temp(sub_prefix: str, pdf_file: UploadFile) -> Path:
    """Save an uploaded file in uploads folder and return its path"""
    file_id = uuid.uuid4().hex
    temp_path = UPLOADS_DIR / f"{sub_prefix}_{file_id}.pdf"
    with temp_path.open("wb") as f:
        shutil.copyfileobj(pdf_file.file, f)
    return temp_path


# --------------------------------------------------------------------
# Type 1 + Type 2 extract endpoints (unchanged logic)
# --------------------------------------------------------------------
@app.post("/extract-type1")
async def extract_type1(pdf_file: UploadFile = File(...)):
    """
    Endpoint for Type 1 PDFs.
    Uses the existing extract_equipment_simple.main(pdf_path, output_path).
    """
    start_time = time.time()

    # Save uploaded PDF
    temp_path = _save_upload_to_temp("type1", pdf_file)

    # Decide where to save Excel (this path is passed into main())
    excel_name = f"equipment_type1_{temp_path.stem}.xlsx"
    excel_path = OUTPUTS_DIR / excel_name

    # IMPORTANT: pass BOTH pdf_path and output_path
    df, equipment_data = run_type1_extractor(str(temp_path), str(excel_path))

    exec_time = round(time.time() - start_time, 2)

    return {
        "message": "Type 1 extraction successful",
        "execution_time_seconds": exec_time,
        "equipment_data": equipment_data,
        "excel_file_name": excel_name,
    }


@app.post("/extract-type2")
async def extract_type2(pdf_file: UploadFile = File(...)):
    """
    Endpoint for Type 2 PDFs.
    Uses run_type2_extractor + save_to_excel from Extract_equipment_simple2.py
    to recreate the same styled Excel you had when running the script manually.
    """
    start_time = time.time()

    # Save uploaded PDF
    temp_path = _save_upload_to_temp("type2", pdf_file)

    # Run new extractor wrapper (returns df + data for UI)
    df, equipment_data = run_type2_extractor(str(temp_path))

    # Figure out "System" name the same way as your manual script
    original_name = pdf_file.filename or "SYSTEM"
    base_name = Path(original_name).name
    system_name = base_name.replace(" Labels.pdf", "").replace(".pdf", "")
    if not system_name.strip():
        system_name = "SYSTEM"

    # Build the input for save_to_excel: {system_name: [rows...]}
    all_systems_data = {system_name: df.to_dict(orient="records")}

    # Save Excel using your original helper
    excel_name = f"equipment_type2_{temp_path.stem}.xlsx"
    excel_path = OUTPUTS_DIR / excel_name
    save_to_excel(all_systems_data, str(excel_path))

    exec_time = round(time.time() - start_time, 2)

    return {
        "message": "Type 2 extraction successful",
        "execution_time_seconds": exec_time,
        "equipment_data": equipment_data,
        "excel_file_name": excel_name,
    }


# --------------------------------------------------------------------
# New: export edited data with styled header
# --------------------------------------------------------------------
class EditedRequest(BaseModel):
    which: str  # "type1" or "type2"
    rows: List[Dict[str, Any]]  # edited rows from the UI


@app.post("/export-edited")
async def export_edited(req: EditedRequest):
    """
    Take edited rows from the UI and generate an Excel file
    with a blue header row (same style as your original files).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Column order must match your styled files
    if req.which.lower() == "type1":
        headers = ["Equipment", "Type", "Properties", "Primary From", "Alternate From"]
    else:
        # Type 2 header order: Alternate From then Primary From (like your SYSTEM A/B files)
        headers = ["Equipment", "Type", "Properties", "Alternate From", "Primary From"]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"

    # Write header row
    ws.append(headers)

    # Style header row (blue fill, white bold text, centered)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row in req.rows:
        ws.append([row.get(h, "") for h in headers])

    # Some reasonable column widths
    ws.column_dimensions["A"].width = 16  # Equipment
    ws.column_dimensions["B"].width = 10  # Type
    ws.column_dimensions["C"].width = 60  # Properties
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Save to outputs/
    file_id = uuid.uuid4().hex
    excel_name = f"edited_{req.which.lower()}_{file_id}.xlsx"
    excel_path = OUTPUTS_DIR / excel_name
    wb.save(excel_path)

    return {"excel_file_name": excel_name}
