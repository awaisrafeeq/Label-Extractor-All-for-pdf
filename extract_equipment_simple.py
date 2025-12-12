"""
Electrical Equipment Data Extractor - Complete with Connection Logic
Extracts SERVICE SWITCHGEAR and DISTRIBUTION SWITCHGEAR with automatic connection mapping
Populates "Primary From" and "Alternate From" based on position and group
"""

import re
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def extract_properties_enhanced(equipment_name, context_text, all_page_text):
    """
    Enhanced property extraction with multiple strategies
    
    Parameters:
    equipment_name (str): Name of equipment (e.g., 'DSG01001')
    context_text (str): Immediate context around equipment
    all_page_text (str): Full page text for extended search
    
    Returns:
    str: Comma-separated properties
    """
    properties = []
    
    # Strategy 1: Look in immediate context
    search_text = context_text
    
    # Strategy 2: If nothing found, search in larger context using equipment name as anchor
    if not any(pattern in context_text.upper() for pattern in ['KVA', 'KV', 'A', 'AMP']):
        # Find the equipment name in full page text and get extended context
        pattern = re.escape(equipment_name)
        match = re.search(pattern, all_page_text)
        if match:
            start = max(0, match.start() - 300)
            end = min(len(all_page_text), match.end() + 300)
            search_text = all_page_text[start:end]
    
    # Extract KVA ratings (e.g., 3350KVA, 2000KVA, 1500KVA)
    kva_matches = re.findall(r'\b(\d+)\s*KVA\b', search_text, re.IGNORECASE)
    if kva_matches:
        # Take the largest KVA value (usually the main rating)
        max_kva = max([int(k) for k in kva_matches])
        properties.append(f"{max_kva}KVA")
    
    # Extract Amperage (e.g., 600A, 1200A, 4000A)
    amp_matches = re.findall(r'\b(\d{3,5})\s*(?:A\b|AMP)', search_text, re.IGNORECASE)
    if amp_matches:
        # Take the largest amperage value
        max_amp = max([int(a) for a in amp_matches])
        properties.append(f"{max_amp}A")
    
    # Extract Primary voltage (e.g., 34.5kV, 13.8kV, 4.16kV)
    primary_volt_matches = re.findall(r'(?:PRIMARY[:\s]+)?(\d+\.?\d*)\s*kV', search_text, re.IGNORECASE)
    if primary_volt_matches:
        # Usually take the first voltage mentioned
        properties.append(f"{primary_volt_matches[0]}kV")
    
    # Extract Secondary voltage (e.g., 480Y/277V, 208Y/120V)
    secondary_volt_match = re.search(r'(?:SECONDARY[:\s]+)?([\d]+Y/[\d]+V)', search_text, re.IGNORECASE)
    if secondary_volt_match:
        properties.append(secondary_volt_match.group(1))
    
    # Extract Voltage ratings in different formats (e.g., 480V, 208V)
    # Only if no secondary voltage found yet
    if not any('Y/' in p for p in properties):
        voltage_matches = re.findall(r'\b(480|208|240|600)\s*V\b', search_text)
        if voltage_matches:
            properties.append(f"{voltage_matches[0]}V")
    
    # DO NOT extract frequency (Hz) or phase information
    # Only extract: Amperage (A), KVA, Primary Voltage (kV), Secondary Voltage (Y/V)
    
    return ', '.join(properties) if properties else ''


def extract_with_positions_pdfplumber(pdf_path):
    """
    Extract equipment with X,Y coordinates using pdfplumber.
    Enhanced version with better property extraction.
    
    Parameters:
    pdf_path: Path to PDF file (str or os.PathLike) or a file-like object.
    
    Returns:
    list: List of equipment dictionaries with positions
    """
    try:
        import pdfplumber
        
        equipment_data = []
        seen_equipment = set()
        
        # Pattern for equipment names
        equipment_pattern = r"'([A-Z]{3}[A-Z0-9]{2}\d{3})'"
        
        # Handle both file path and file-like object
        if isinstance(pdf_path, (str, os.PathLike)):
            pdf_obj = pdfplumber.open(pdf_path)
        else:
            # assume file-like, reset to start
            try:
                pdf_path.seek(0)
            except Exception:
                pass
            pdf_obj = pdfplumber.open(pdf_path)
        
        with pdf_obj as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Get full page text for extended context searches
                full_page_text = page.extract_text() or ""
                
                # Extract words with positions
                words = page.extract_words(x_tolerance=3, y_tolerance=3)
                
                # Create text blocks grouped by proximity
                for i, word in enumerate(words):
                    text = word['text']
                    
                    # Check if this word contains equipment name
                    match = re.search(equipment_pattern, text)
                    if match:
                        equipment_name = match.group(1)
                        equipment_type = equipment_name[:3]
                        
                        # Filter for MVS and DSG only
                        if equipment_type not in ['MVS', 'DSG']:
                            continue
                        
                        # Skip duplicates
                        if equipment_name in seen_equipment:
                            continue
                        seen_equipment.add(equipment_name)
                        
                        # Get surrounding text for properties (larger window)
                        context_start = max(0, i - 10)
                        context_end = min(len(words), i + 40)
                        context_text = ' '.join([w['text'] for w in words[context_start:context_end]])
                        
                        # Extract properties with enhanced method
                        properties = extract_properties_enhanced(equipment_name, context_text, full_page_text)
                        
                        equipment_data.append({
                            'Equipment': equipment_name,
                            'Type': equipment_type,
                            'Properties': properties,
                            'Alternate From': '',
                            'Primary From': '',
                            'x_position': word['x0'],  # Left X coordinate
                            'y_position': word['top'],  # Top Y coordinate
                            'page': page_num
                        })
        
        # Sort by page, then Y (top to bottom), then X (left to right)
        equipment_data.sort(key=lambda x: (x['page'], x['y_position'], x['x_position']))
        
        return equipment_data
        
    except Exception as e:
        print(f"pdfplumber extraction failed: {e}")
        return None


def extract_with_positions_pypdf2(pdf_path):
    """
    Extract equipment with approximate positions using PyPDF2.
    Enhanced version with better property extraction.
    
    Parameters:
    pdf_path: Path to PDF file (str or os.PathLike) or a file-like object.
    
    Returns:
    list: List of equipment dictionaries
    """
    try:
        import PyPDF2
        
        equipment_data = []
        seen_equipment = set()
        
        equipment_pattern = r"'([A-Z]{3}[A-Z0-9]{2}\d{3})'"
        
        # Handle both file path and file-like object
        if isinstance(pdf_path, (str, os.PathLike)):
            file_obj = open(pdf_path, 'rb')
            close_after = True
        else:
            file_obj = pdf_path
            close_after = False
            try:
                file_obj.seek(0)
            except Exception:
                pass
        
        try:
            pdf_reader = PyPDF2.PdfReader(file_obj)
            
            for page_num, page in enumerate(pdf_reader.pages):
                full_page_text = page.extract_text() or ""
                if not full_page_text:
                    continue
                
                # Split into lines and process
                lines = full_page_text.split('\n')
                
                for line_num, line in enumerate(lines):
                    matches = re.finditer(equipment_pattern, line)
                    
                    for match in matches:
                        equipment_name = match.group(1)
                        equipment_type = equipment_name[:3]
                        
                        # Filter for MVS and DSG only
                        if equipment_type not in ['MVS', 'DSG']:
                            continue
                        
                        # Skip duplicates
                        if equipment_name in seen_equipment:
                            continue
                        seen_equipment.add(equipment_name)
                        
                        # Get context from surrounding lines (larger window)
                        start_line = max(0, line_num - 5)
                        end_line = min(len(lines), line_num + 10)
                        context_text = ' '.join(lines[start_line:end_line])
                        
                        # Extract properties with enhanced method
                        properties = extract_properties_enhanced(equipment_name, context_text, full_page_text)
                        
                        equipment_data.append({
                            'Equipment': equipment_name,
                            'Type': equipment_type,
                            'Properties': properties,
                            'Alternate From': '',
                            'Primary From': '',
                            'x_position': match.start(),  # Approximate position in line
                            'y_position': line_num,
                            'page': page_num
                        })
        finally:
            if close_after:
                file_obj.close()
        
        # Sort by page, line, and position in line
        equipment_data.sort(key=lambda x: (x['page'], x['y_position'], x['x_position']))
        
        return equipment_data
        
    except Exception as e:
        print(f"PyPDF2 extraction failed: {e}")
        return None


def identify_dsg_groups(equipment_data):
    """
    Identify DSG groups based on their naming pattern and Y-position.
    DSGs with same 5th character (group letter) and similar Y-position form a group.
    
    Parameters:
    equipment_data (list): List of equipment dictionaries
    
    Returns:
    dict: Dictionary with group identifiers as keys and lists of DSG items as values
    """
    
    dsg_items = [item for item in equipment_data if item['Type'] == 'DSG']
    
    # Group DSGs by their 4th and 5th characters (system identifier) and page
    # e.g., DSGAA110 -> 'AA', DSGCA110 -> 'CA'
    groups = {}
    
    for item in dsg_items:
        equipment_name = item['Equipment']
        page = item['page']
        
        # Extract group identifier (4th character)
        group_letter = equipment_name[3]  # 4th character: A, B, C, D, E, F
        
        # Create unique group key with page number
        group_key = f"Page{page+1}_Group{group_letter}"
        
        if group_key not in groups:
            groups[group_key] = []
        
        groups[group_key].append(item)
    
    # Sort each group by x_position (left to right)
    for group_key in groups:
        groups[group_key].sort(key=lambda x: x['x_position'])
    
    return groups


def populate_connections(equipment_data):
    """
    Populate Primary From and Alternate From for all DSG equipment.
    
    Logic:
    - Leftmost DSG: Primary = Leftmost MVS, Alternate = Next DSG (right)
    - Middle DSGs: Primary = Previous DSG (left), Alternate = Next DSG (right)
    - Rightmost DSG: Primary = Rightmost MVS, Alternate = Previous DSG (left)
    
    Parameters:
    equipment_data (list): List of equipment dictionaries
    
    Returns:
    list: Updated equipment data with connections populated
    """
    
    # Separate MVS and DSG items
    mvs_items = [item for item in equipment_data if item['Type'] == 'MVS']
    
    # Group MVS by page
    mvs_by_page = {}
    for item in mvs_items:
        page = item['page']
        if page not in mvs_by_page:
            mvs_by_page[page] = []
        mvs_by_page[page].append(item)
    
    # Sort MVS items by x_position for each page
    for page in mvs_by_page:
        mvs_by_page[page].sort(key=lambda x: x['x_position'])
    
    # Identify DSG groups
    dsg_groups = identify_dsg_groups(equipment_data)
    
    print("\n" + "="*70)
    print("CONNECTION MAPPING")
    print("="*70)
    
    # Process each DSG group
    for group_key, dsg_list in dsg_groups.items():
        print(f"\n{group_key}: {len(dsg_list)} DSGs")
        
        if len(dsg_list) == 0:
            continue
        
        # Get page number for this group
        page = dsg_list[0]['page']
        
        # Get MVS items for this page
        page_mvs = mvs_by_page.get(page, [])
        
        if len(page_mvs) < 2:
            print(f"  ⚠️  Warning: Less than 2 MVS items found on page {page+1}")
            continue
        
        leftmost_mvs = page_mvs[0]['Equipment']
        rightmost_mvs = page_mvs[-1]['Equipment']
        
        # Process each DSG in the group
        for i, dsg in enumerate(dsg_list):
            if i == 0:
                # Leftmost DSG
                dsg['Primary From'] = leftmost_mvs
                if len(dsg_list) > 1:
                    dsg['Alternate From'] = dsg_list[i + 1]['Equipment']
                print(f"  {dsg['Equipment']}: Primary={leftmost_mvs}, Alternate={dsg['Alternate From']}")
                
            elif i == len(dsg_list) - 1:
                # Rightmost DSG
                dsg['Primary From'] = rightmost_mvs
                dsg['Alternate From'] = dsg_list[i - 1]['Equipment']
                print(f"  {dsg['Equipment']}: Primary={rightmost_mvs}, Alternate={dsg['Alternate From']}")
                
            else:
                # Middle DSGs
                dsg['Primary From'] = dsg_list[i - 1]['Equipment']
                dsg['Alternate From'] = dsg_list[i + 1]['Equipment']
                print(f"  {dsg['Equipment']}: Primary={dsg['Primary From']}, Alternate={dsg['Alternate From']}")
    
    return equipment_data


def extract_from_pdf(pdf_path):
    """
    Extract equipment data from PDF file in left-to-right order.
    Tries pdfplumber first (best for coordinates), then PyPDF2.
    
    Parameters:
    pdf_path: Path to input PDF (str or os.PathLike) or a file-like object.
    
    Returns:
    list: List of equipment dictionaries sorted by position
    """
    
    print("Attempting coordinate-based extraction with pdfplumber...")
    equipment_data = extract_with_positions_pdfplumber(pdf_path)
    
    if equipment_data is not None and len(equipment_data) > 0:
        print(f"✓ Successfully extracted {len(equipment_data)} items with pdfplumber")
        return equipment_data
    
    # If pdf_path is a file-like object, reset pointer before fallback
    try:
        if not isinstance(pdf_path, (str, os.PathLike)):
            pdf_path.seek(0)
    except Exception:
        pass
    
    print("Falling back to PyPDF2...")
    equipment_data = extract_with_positions_pypdf2(pdf_path)
    
    if equipment_data is not None and len(equipment_data) > 0:
        print(f"✓ Successfully extracted {len(equipment_data)} items with PyPDF2")
        return equipment_data
    
    return None


def create_excel_file(equipment_data, output_path):
    """
    Create formatted Excel file from equipment data.
    
    Parameters:
    equipment_data (list): List of equipment dictionaries
    output_path (str): Output file path
    """
    
    # Remove position data before creating DataFrame
    clean_data = []
    for item in equipment_data:
        clean_data.append({
            'Equipment': item['Equipment'],
            'Type': item['Type'],
            'Properties': item['Properties'],
            'Alternate From': item['Alternate From'],
            'Primary From': item['Primary From']
        })
    
    # Create DataFrame (already sorted by position)
    df = pd.DataFrame(clean_data)
    
    # Create Excel workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Equipment Data'
    
    # Headers
    headers = ['Equipment', 'Type', 'Properties', 'Alternate From', 'Primary From']
    sheet.append(headers)
    
    # Format headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num in range(1, 6):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row_data in clean_data:
        sheet.append([
            row_data['Equipment'],
            row_data['Type'],
            row_data['Properties'],
            row_data['Alternate From'],
            row_data['Primary From']
        ])
    
    # Column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    
    # Borders and alignment
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Save
    wb.save(output_path)
    
    return df


def print_summary(df, equipment_data):
    """Print summary of extracted data with connection info"""
    print("\n" + "="*70)
    print("EXTRACTION SUMMARY")
    print("="*70)
    
    print(f"\n✓ Total records: {len(df)}")
    print(f"\nEquipment breakdown:")
    
    type_names = {
        'DSG': 'Distribution Switchgear',
        'MVS': 'Service Switchgear'
    }
    
    for eq_type in sorted(df['Type'].unique()):
        count = len(df[df['Type'] == eq_type])
        type_name = type_names.get(eq_type, 'Unknown')
        
        # Count how many have properties
        with_props = len([item for item in equipment_data if item['Type'] == eq_type and item['Properties']])
        
        # Count how many DSGs have connections
        if eq_type == 'DSG':
            with_connections = len([item for item in equipment_data 
                                   if item['Type'] == eq_type 
                                   and (item['Primary From'] or item['Alternate From'])])
            print(f"  {eq_type} ({type_name}): {count} items ({with_props} with properties, {with_connections} with connections)")
        else:
            print(f"  {eq_type} ({type_name}): {count} items ({with_props} with properties)")
    
    # Check for items without properties
    no_props = [item for item in equipment_data if not item['Properties']]
    if no_props:
        print(f"\n⚠️  Warning: {len(no_props)} items found without properties:")
        for item in no_props:
            print(f"   - {item['Equipment']}")
    
    # Check for DSGs without connections
    dsg_no_connections = [item for item in equipment_data 
                          if item['Type'] == 'DSG' 
                          and not (item['Primary From'] or item['Alternate From'])]
    if dsg_no_connections:
        print(f"\n⚠️  Warning: {len(dsg_no_connections)} DSGs found without connections:")
        for item in dsg_no_connections:
            print(f"   - {item['Equipment']}")


def main(pdf_path, output_path):
    """
    Main extraction function.
    Extracts SERVICE SWITCHGEAR (MVS) and DISTRIBUTION SWITCHGEAR (DSG) in left-to-right order.
    Automatically populates Primary From and Alternate From connections.
    
    Parameters:
    pdf_path: Path to input PDF (str or os.PathLike) or a file-like object.
    output_path (str): Path to output Excel file
    """
    
    print("=" * 70)
    print("Electrical Equipment Data Extractor - With Connection Mapping")
    print("Extracts MVS & DSG with automatic Primary/Alternate connections")
    print("=" * 70)
    print(f"\nInput PDF: {pdf_path}")
    print(f"Output Excel: {output_path}\n")
    
    # Extract data with positions
    equipment_data = extract_from_pdf(pdf_path)
    
    if equipment_data is None or len(equipment_data) == 0:
        print("\n✗ No MVS or DSG equipment found in PDF")
        return None
    
    # Populate connections for DSG items
    equipment_data = populate_connections(equipment_data)
    
    # Create Excel file
    df = create_excel_file(equipment_data, output_path)
    print(f"\n✓ Excel file created: {output_path}")
    
    # Print summary
    print_summary(df, equipment_data)
    
    print("\n" + "=" * 70)
    print("Extraction completed successfully!")
    print("=" * 70)
    
    return df, equipment_data


if __name__ == "__main__":
    # Usage example with file path
    PDF_PATH = r'E:\Ai Data House Intern\Austin-hurt-label Extractor\Code\newcode-file\MEDIUM VOLTAGE.pdf'
    OUTPUT_PATH = r'E:\Ai Data House Intern\Austin-hurt-label Extractor\Code\newcode-file\equipment_data_complete.xlsx'
    
    df, equipment_data = main(PDF_PATH, OUTPUT_PATH)
